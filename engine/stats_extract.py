#!/usr/bin/env python3
"""
지자체 통계연보(엑셀) → 지역개황(0200) 표 값 추출.

`calc.py`(소음진동) · `calc_air.py`(대기질)와 같은 자리의 파트 모듈이다. 다만 성질이 다르다 —
**계산이 아니라 소싱**이다. 지역개황 값의 대부분은 공식으로 유도되지 않고 통계연보에서 옮겨진다
(rule `small-env/regional-overview.md` §2).

  통계연보 zip/폴더  →  편(xlsx)  →  시트  →  연도 행/지역 행  →  단위 변환  →  표 값

검증: `python engine/stats_extract.py --self-test`
      천안 화덕리 골든셋(§2.2.1 지목별 · §2.5.4 자동차)과 대조한다.

⚠️ 이 모듈은 **엑셀 배포판만** 다룬다. 스캔 PDF 판(평창 등)은 OCR 없이는 값을 꺼낼 수 없다.
   배포 형식별 현황은 `catalog/review/stats_catalog.md` §2.
"""
import argparse, io, re, sys, zipfile
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl 이 필요합니다: .venv/bin/pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent

# ── 절 → (편 패턴, 시트 패턴) 매핑 ────────────────────────────────────────────
# 통계연보 편 구성은 통계청 표준이라 지자체가 달라도 같다 (소싱실증 §3-1).
# ✅ = 골든셋으로 검증됨 / ⬜ = 매핑만 확인, 파서 미구현
SHEET_MAP = {
    "2.2.1 지목별 토지이용":      ("02.토지", r"토지\s*지목별",          "✅"),
    "2.5.4 자동차":               ("11.교통", r"자동차\s*등록$|1-자동차등록", "✅"),
    "2.2.2 용도지역":             ("10.주택", r"용도지역",               "✅"),
    "2.5.1 도로":                 ("10.주택", r"^\d+-도로|도\s*로$",      "✅"),
    "2.5.2 환경오염물질 배출시설": ("13. 환",  r"환경오염물질\s*배출사업장", "✅"),
    "2.6.3 문화재":               ("14-02",   r"문화재",                 "✅"),
}

# 지목 이름 — 통계연보 헤더 표기에는 균등분할 공백이 들어간다 (`과 수 원`).
JIMOK = ["합계", "전", "답", "과수원", "목장용지", "임야", "광천지", "대", "공장용지",
         "학교용지", "주차장", "주유소용지", "창고용지", "도로", "철도용지", "제방",
         "하천", "구거", "유지", "양어장", "수도용지", "공원", "체육용지", "유원지",
         "종교용지", "사적지", "묘지", "잡종지"]


def _norm(s):
    """균등분할 공백·개행·괄호주석을 지운 비교용 문자열."""
    if s is None:
        return ""
    return re.sub(r"[\s ]+", "", str(s)).replace("·", "")


class YearBook:
    """통계연보 한 판(版). zip 이든 풀린 폴더든 같게 다룬다."""

    def __init__(self, path):
        self.path = Path(path)
        self._books = {}                      # 편 파일명 → bytes
        if self.path.is_dir():
            for p in sorted(self.path.glob("*.xls*")):
                self._books[p.name] = p.read_bytes()
        elif self.path.suffix.lower() == ".zip":
            z = zipfile.ZipFile(self.path)
            for i in z.infolist():
                if i.is_dir():
                    continue
                # ⚠️ 지자체 배포 zip 은 파일명이 CP949 다. unzip 은 여기서 실패한다.
                try:
                    name = i.filename.encode("cp437").decode("cp949")
                except Exception:
                    name = i.filename
                if name.lower().endswith((".xlsx", ".xls")):
                    self._books[Path(name).name] = z.read(i)
        else:
            self._books[self.path.name] = self.path.read_bytes()
        if not self._books:
            raise ValueError(f"엑셀 편을 찾지 못했습니다: {path}")

    @property
    def volumes(self):
        return sorted(self._books)

    def sheet(self, vol_pat, sheet_pat):
        """편 패턴·시트 패턴으로 워크시트를 찾는다. 이름이 지자체마다 조금씩 다르므로 정규식."""
        for name, blob in self._books.items():
            if not re.search(vol_pat, name):
                continue
            wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True)
            for sn in wb.sheetnames:
                if re.search(sheet_pat, sn):
                    return wb[sn]
        return None


# ── 표 읽기 공통 ────────────────────────────────────────────────────────────
def _header_row(ws, must_have="합계", limit=12):
    """헤더 행 번호를 찾는다. 통계연보는 제목·단위 줄이 앞에 붙어 행 번호가 판마다 다르다."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=limit, values_only=True), 1):
        # 헤더 셀에 영문이 붙어 있는 판이 많다 (`합계Total`) — 앞부분만 본다.
        if any(_norm(c).startswith(must_have) for c in row):
            return i
    return None


def _year_rows(ws, label_col=1, max_row=80):
    """라벨 열에서 `2022` 꼴 연도 행을 모은다 → {연도: 행번호}.

    ⚠️ 엑셀이 연도를 숫자로 들고 있어 `2,018` 처럼 보이는 판도 있다.
    ⚠️ 연도 블록 사이에 **빈 행**이 끼는 판이 있다 — 마지막 행을 놓치기 쉬운 구조이고,
       실제로 정답지가 이 함정에 빠져 있었다 (소싱실증 §4-3)."""
    out = {}
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, values_only=True), 1):
        v = row[label_col - 1]
        if v is None:
            continue
        m = re.fullmatch(r"(19|20)\d{2}", _norm(v).replace(",", ""))
        if m:
            out[int(m.group(0))] = i
    return out


def _region_row(ws, region, label_col=1, max_row=400):
    """읍면동 블록에서 지역 행을 찾는다. `동    면` 처럼 균등분할 공백이 들어간다."""
    tgt = _norm(region)
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, values_only=True), 1):
        if _norm(row[label_col - 1]) == tgt:
            return i
    return None


def _blocks(ws, row, skip=("Year", "연별", "읍면동")):
    """상위 헤더 행을 **블록**으로 쪼갠다 → [(이름, 시작열, 끝열)].

    통계연보 표는 위 칸이 큰 묶음(대기 · 수질), 아래 칸이 같은 이름의 반복(계 · 1종 …)이다.
    이름만으로 찾으면 대기의 `계` 와 수질의 `계` 가 구분되지 않는다.
    ⚠️ 표가 옆으로 길면 중간에 `Year`·`연별` 이 다시 끼어든다 — 걸러야 블록이 안 어긋난다."""
    marks = []
    for c in range(1, ws.max_column + 1):
        v = _norm(ws.cell(row, c).value)
        if v and not any(k in v for k in skip):
            marks.append((c, v))
    return [(v, c, (marks[i + 1][0] - 1 if i + 1 < len(marks) else ws.max_column))
            for i, (c, v) in enumerate(marks)]


def _col_in(ws, rows, lo, hi, key):
    """블록 [lo,hi] 안에서 라벨이 `key` 로 시작하는 첫 열.

    ⚠️ **행을 먼저** 훑는다 — 열을 먼저 훑으면 아래층의 비슷한 이름이 먼저 걸린다.
       용도지역에서 `도시지역`(면적)을 찾다가 `도시지역인구1)` 를 잡았다."""
    for r in rows:
        for c in range(lo, hi + 1):
            if _norm(ws.cell(r, c).value).startswith(key):
                return c
    return None


def _row_values(ws, r):
    return list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]


def _pick_row(ws, year=None, region=None):
    """연도 행 또는 읍면동 행 → (행번호, 무엇을 썼는지)."""
    if region:
        r = _region_row(ws, region)
        if r is None:
            raise LookupError(f"'{region}' 행을 찾지 못했습니다")
        return r, f"{region} 행"
    yrs = _year_rows(ws)
    if not yrs:
        raise LookupError("연도 행을 찾지 못했습니다")
    y = year or max(yrs)
    if y not in yrs:
        raise LookupError(f"{y}년 행이 없습니다 (있는 해: {sorted(yrs)})")
    return yrs[y], f"{y}년 행"


# ── 2.2.2 용도지역 ──────────────────────────────────────────────────────────
# 보고서 표는 도시지역 4종 · 비도시지역 3종으로 접는다.
# ⚠️ 통계연보의 **관리지역은 계획·생산·보전 셋으로 나뉘어 있다** — 보고서는 이를 합쳐 쓴다.
ZONING = {
    "주거": ["주거지역"], "상업": ["상업지역"], "공업": ["공업지역"], "녹지": ["녹지지역"],
    "관리": ["계획관리지역", "생산관리지역", "보전관리지역"],
    "농림": ["농림지역"], "보전": ["자연환경보전지역"],
}


def zoning(yb, year=None):
    """용도지역 면적을 ㎢ 로. 통계연보 단위가 **천㎡** 라 1,000 으로 나눈다."""
    ws = yb.sheet(*SHEET_MAP["2.2.2 용도지역"][:2])
    if ws is None:
        raise LookupError("용도지역 시트를 찾지 못했습니다")
    rows = range(6, 10)
    hi = ws.max_column
    col = {}
    for k, names in ZONING.items():
        col[k] = [_col_in(ws, rows, 1, hi, n) for n in names]
    tot = _col_in(ws, rows, 1, hi, "용도지역총합계") or _col_in(ws, rows, 1, hi, "총합계")
    urban = _col_in(ws, rows, 1, hi, "도시지역")
    nonurban = _col_in(ws, rows, 1, hi, "계(A)")

    r, used = _pick_row(ws, year)
    v = _row_values(ws, r)

    def g(c):
        x = v[c - 1] if c and c <= len(v) else None
        return x if isinstance(x, (int, float)) else 0

    total, city = round(g(tot) / 1000, 2), round(g(urban) / 1000, 2)
    # 비도시는 **빼서** 낸다 — 직접 읽으면 반올림이 어긋나 합계와 안 맞는다
    # (천안 2021: 직접 492.13 ↔ 빼기 492.14, 정답은 후자다)
    out = {"합계": total, "도시지역계": city, "비도시지역계": round(total - city, 2)}
    for k, cs in col.items():
        out[k] = round(sum(g(c) for c in cs) / 1000, 2)
    out["_출처행"] = used
    return out


# ── 2.5.1 도로 ─────────────────────────────────────────────────────────────
ROAD_KINDS = ["고속도로", "일반국도", "지방도", "시군도"]


def roads(yb, year=None):
    """도로 종류별 개통연장·포장·미포장·미개통·포장률."""
    ws = yb.sheet(*SHEET_MAP["2.5.1 도로"][:2])
    if ws is None:
        raise LookupError("도로 시트를 찾지 못했습니다")
    rows = range(7, 10)
    blocks = {n: (lo, hi) for n, lo, hi in _blocks(ws, 6)}
    r, used = _pick_row(ws, year)
    v = _row_values(ws, r)

    def g(c):
        x = v[c - 1] if c and c <= len(v) else None
        return x if isinstance(x, (int, float)) else None

    out = {}
    for kind in ROAD_KINDS + ["합계"]:
        key = next((n for n in blocks if n.startswith(kind)), None)
        if not key:
            continue
        lo, hi = blocks[key]
        # 합계 블록만 상위 칸에 종류 이름이 없어 `개통` 대신 표 머리를 그대로 쓴다
        row = {"개통연장": g(_col_in(ws, rows, lo, hi, "개통") or lo),
               "포장": g(_col_in(ws, rows, lo, hi, "포장")),
               "미포장": g(_col_in(ws, rows, lo, hi, "미포장")),
               "미개통": g(_col_in(ws, rows, lo, hi, "미개통")),
               "포장률": g(_col_in(ws, rows, lo, hi, "포장률"))}
        if row["포장률"] is not None:
            row["포장률"] = round(row["포장률"], 1)
        out[kind] = row
    out["_출처행"] = used
    return out


# ── 2.5.2 환경오염물질 배출시설 ─────────────────────────────────────────────
def emitters(yb, year=None):
    """대기·수질 종별 배출사업장 수와 소음·진동."""
    ws = yb.sheet(*SHEET_MAP["2.5.2 환경오염물질 배출시설"][:2])
    if ws is None:
        raise LookupError("환경오염물질 배출사업장 시트를 찾지 못했습니다")
    rows = range(7, 9)
    blocks = {n: (lo, hi) for n, lo, hi in _blocks(ws, 6)}
    r, used = _pick_row(ws, year)
    v = _row_values(ws, r)

    def g(c):
        x = v[c - 1] if c and c <= len(v) else None
        return x if isinstance(x, (int, float)) else None

    out = {}
    for label, head in (("대기", "대기"), ("수질", "수질")):
        key = next((n for n in blocks if n.startswith(head)), None)
        if not key:
            continue
        lo, hi = blocks[key]
        out[label] = {"계": g(_col_in(ws, rows, lo, hi, "계"))}
        for n in range(1, 6):
            out[label][f"{n}종"] = g(_col_in(ws, rows, lo, hi, f"{n}종"))
    key = next((n for n in blocks if n.startswith("소음")), None)
    if key:
        lo, hi = blocks[key]
        # 보고서는 소음·진동을 **한 칸으로 합쳐** 적는다 (천안 1,324 = 소음 + 진동)
        a, b = g(_col_in(ws, rows, lo, hi, "소음")), g(_col_in(ws, rows, lo, hi, "진동"))
        out["소음진동"] = (a or 0) + (b or 0)
    out["_출처행"] = used
    return out


# ── 2.6.3 문화재 ───────────────────────────────────────────────────────────
HERITAGE = ["국보", "보물", "사적", "명승", "천연기념물", "국가무형문화재",
            "국가민속문화재", "시도유형문화재", "시도무형문화재", "시도기념물",
            "시도민속문화재", "문화재자료", "국가등록문화재", "시도등록문화재"]


def heritage(yb, year=None, region=None):
    """문화재 지정현황. `region` 을 주면 읍면동 행 — 사업지 소재 면에 지정이 있는지 본다."""
    ws = yb.sheet(*SHEET_MAP["2.6.3 문화재"][:2])
    if ws is None:
        raise LookupError("문화재 시트를 찾지 못했습니다")
    rows = range(6, 9)
    hi = ws.max_column
    r, used = _pick_row(ws, year, region)
    v = _row_values(ws, r)

    def g(c):
        x = v[c - 1] if c and c <= len(v) else None
        return x if isinstance(x, (int, float)) else 0

    out = {"총계": g(_col_in(ws, rows, 1, hi, "총계"))}
    for k in HERITAGE:
        out[k] = g(_col_in(ws, rows, 1, hi, k))
    # 보고서 본문은 묶음 단위로 쓴다 — "국가지정 14, 시도지정 27, 문화재자료 25"
    out["국가지정계"] = sum(out[k] for k in HERITAGE[:7])
    out["시도지정계"] = sum(out[k] for k in HERITAGE[7:11])
    out["_출처행"] = used
    return out


# ── 2.2.1 지목별 토지이용 ───────────────────────────────────────────────────
def land_use(yb, region=None, year=None):
    """지목별 면적을 ㎢ 로 돌려준다.

    region=None 이면 연도 행(시군 전체), region 을 주면 읍면동 블록의 그 행.
    year=None 이면 **가장 최근 연도 행** — 어느 해를 쓸지는 회사 표준 확인 대상
    (`docs/실무자_확인요청.md` F-1)."""
    ws = yb.sheet(*SHEET_MAP["2.2.1 지목별 토지이용"][:2])
    if ws is None:
        raise LookupError("토지 지목별 현황 시트를 찾지 못했습니다")

    hr = _header_row(ws)
    if hr is None:
        raise LookupError("지목 헤더 행을 찾지 못했습니다")
    header = list(ws.iter_rows(min_row=hr, max_row=hr, values_only=True))[0]
    col = {}
    for ci, cell in enumerate(header, 1):
        n = _norm(cell)
        if n in JIMOK and n not in col:          # 블록이 반복되므로 처음 것만
            col[n] = ci

    if region:
        r = _region_row(ws, region)
        if r is None:
            raise LookupError(f"'{region}' 행을 찾지 못했습니다")
        used = f"{region} 행"
    else:
        yrs = _year_rows(ws)
        if not yrs:
            raise LookupError("연도 행을 찾지 못했습니다")
        y = year or max(yrs)
        r = yrs[y]
        used = f"{y}년 행"

    vals = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    out = {}
    for k, ci in col.items():
        v = vals[ci - 1]
        if isinstance(v, (int, float)):
            out[k] = round(v / 1_000_000, 2)      # ㎡ → ㎢, 소수 2자리 (실증 §4-1)
    out["_출처행"] = used
    return out


def land_use_ratio(area: dict):
    """구성비(%) — rule §3-1. 면적÷합계×100, 소수 2자리."""
    tot = area.get("합계")
    if not tot:
        return {}
    return {k: round(v / tot * 100, 2) for k, v in area.items()
            if isinstance(v, (int, float))}


# ── 2.5.4 자동차 ────────────────────────────────────────────────────────────
def vehicles(yb, year=None):
    """자동차 등록대수. 보고서 표는 계·승용차·승합차·화물차·특수차·이륜자동차 6칸이다."""
    ws = yb.sheet(*SHEET_MAP["2.5.4 자동차"][:2])
    if ws is None:
        raise LookupError("자동차 등록 시트를 찾지 못했습니다")
    hr = _header_row(ws)
    header = list(ws.iter_rows(min_row=hr, max_row=hr, values_only=True))[0] if hr else ()
    col = {}
    for ci, cell in enumerate(header, 1):
        n = _norm(cell)
        for k in ("합계", "승용차", "승합차", "화물차", "특수차", "이륜자동차"):
            if n.startswith(k) and k not in col:
                col[k] = ci
    yrs = _year_rows(ws)
    y = year or max(yrs)
    vals = list(ws.iter_rows(min_row=yrs[y], max_row=yrs[y], values_only=True))[0]
    out = {k: vals[ci - 1] for k, ci in col.items() if isinstance(vals[ci - 1], (int, float))}
    out["_출처행"] = f"{y}년 행"
    return out


# ── 자체 검증 ───────────────────────────────────────────────────────────────
# 천안 화덕리 골든셋 §2.2.1 · §2.5.4 실측값 (golden/small-env/천안_화덕리/regional-overview.txt)
GOLDEN_CHEONAN = {
    "시군_2021": {"합계": 636.15, "전": 51.88, "답": 92.85, "임야": 309.02,
                  "대": 39.69, "과수원": 15.68},
    "동면":      {"합계": 43.27, "전": 4.57, "답": 6.23, "임야": 26.18,
                  "대": 0.86, "과수원": 0.09},
    "자동차":     {"합계": 351796, "승용차": 294246, "승합차": 10499,
                  "화물차": 45382, "특수차": 1669, "이륜자동차": 22622},
    # 용도지역도 2021년 행이다 (토지이용과 같은 해)
    "용도지역_2021": {"합계": 636.08, "도시지역계": 143.94, "비도시지역계": 492.14,
                     "주거": 35.03, "상업": 3.25, "공업": 16.16, "녹지": 89.50,
                     "관리": 230.11, "농림": 260.17},
    # 도로·배출시설·문화재는 2022년 기준이다 (출처 주석에 그렇게 적혀 있다)
    "도로_2022":   {"고속도로": 51410, "일반국도": 120162, "지방도": 101629,
                   "시군도": 1425958, "합계": 1699159},
    "배출_2022":   {"대기": 1033, "수질": 1070, "소음진동": 1324},
    "문화재_2022": {"총계": 102, "국가지정계": 14, "시도지정계": 27, "문화재자료": 25},
}


# ═══ 0500 환경현황 5.3 — 인구·주택·사업체 (build_env_status_vars.py 가 부른다) ═══
# ⚠️ 지역개황(0200)에는 인구 절이 없다 — 이 표들은 0500 의 5.3 에만 쓰인다.
#    같은 통계연보 원자료를 쓰므로 이 모듈에 둔다 (원천 단일화).
#    검증: `--self-test-0500` — 원주 2024 기본통계(웹 재배포판) vs 원주 env-status 골든.

def _numc(v):
    """통계연보 셀 → 숫자. '35,145' 같은 콤마 문자열 판이 있다 (원주 2022 사업체 행 실측).
    '…'(자료없음)·'-'·빈칸은 None."""
    if v is None:
        return None
    s = str(v).strip().replace(",", "")
    if s in ("", "…", "-", "···"):
        return None
    try:
        return int(s) if re.fullmatch(r"-?\d+", s) else float(s)
    except ValueError:
        return None


def _year_rows_loose(ws, label_col=1, max_row=300):
    """연도 행 — `20221)` 처럼 **각주 번호가 붙은 판**이 있다 (원주 사업체 총괄 실측).
    `_year_rows` 의 fullmatch 는 그 행을 놓친다 — 접두 매칭으로 줍는다."""
    out = {}
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, values_only=True), 1):
        v = row[label_col - 1]
        if v is None:
            continue
        m = re.match(r"((?:19|20)\d{2})(?:\d?\))?$", _norm(v).replace(",", ""))
        if m:
            out[int(m.group(1))] = i
    return out


def _find_col(ws, hdr_rows, key, lo=1, hi=None):
    """헤더 행 구간에서 라벨이 key 로 시작하는 열. 행 우선 (`도시지역인구1)` 사고 방지)."""
    hi = hi or ws.max_column
    for r in hdr_rows:
        for c in range(lo, hi + 1):
            if _norm(ws.cell(r, c).value).startswith(key):
                return c
    return None


def 인구_연별(yb, n=5):
    """읍면동별 세대·인구 시트의 **연별(시 전체) 행** → [[연, 세대, 인구, 남, 여, 세대당]…].

    등록인구 **합계**(한국인+외국인) 기준 — 골든 5.3 인구추이가 이 열이다 (원주 2023
    366,279 = 한국인 361,503 + 외국인, 실측). 세대당은 시트에 없어 계산한다 (표시 1자리).
    ⚠️ 이 시트에 `=SUM` 수식이 살아 있는 판이 있다 — data_only 캐시로 읽힌다 (원주 2020~21)."""
    ws = yb.sheet(r"인구", r"읍.?면.?동별\s*세대")
    if ws is None:
        return None
    yrs = _year_rows(ws)
    out = []
    for y in sorted(yrs)[-n:]:
        r = _row_values(ws, yrs[y])
        세대, 인구, 남, 여 = (_numc(r[1]), _numc(r[2]), _numc(r[3]), _numc(r[4]))
        세대당 = round(인구 / 세대, 1) if 인구 and 세대 else None
        out.append([y, 세대, 인구, 남, 여, 세대당])
    return out or None


def 인구_읍면(yb, 읍면):
    """최신 연도 블록의 읍면동 행 → {세대, 인구, 남, 여}. (연별 행 아래 붙어 있다)"""
    ws = yb.sheet(r"인구", r"읍.?면.?동별\s*세대")
    if ws is None:
        return None
    r = _region_row(ws, 읍면)
    if r is None:
        return None
    v = _row_values(ws, r)
    return {"세대": _numc(v[1]), "인구": _numc(v[2]), "남": _numc(v[3]), "여": _numc(v[4])}


def 인구동태(yb, n=5):
    """인구동태(출생·사망) + 인구이동(전입·전출) 두 시트 → [[연, 출생, 사망, 전입, 전출]…].

    순증감(▲▼)은 여기서 계산하지 않는다 — 조립(핸들러)이 표시 계층이다.
    ⚠️ 연도 행 뒤에 월별 행(`1월`…)이 붙는다 — `_year_rows` 가 연도만 줍는다."""
    d = yb.sheet(r"인구", r"인구동태$|^\d[\s.]*인구동태")
    m = yb.sheet(r"인구", r"인구이동")
    if d is None or m is None:
        return None
    yd, ym = _year_rows(d), _year_rows(m)
    out = []
    for y in sorted(set(yd) & set(ym))[-n:]:
        rd, rm = _row_values(d, yd[y]), _row_values(m, ym[y])
        out.append([y, _numc(rd[1]), _numc(rd[4]), _numc(rm[1]), _numc(rm[4])])
    return out or None


def 주택현황(yb, n=5):
    """주택현황·보급률 → [[연, 가구, 합계, 단독, 아파트, 연립, 다세대, 비거주, 보급률]…].

    열은 라벨로 찾는다 — 단독주택 아래 `다가구주택` 내수 열이 끼어 있어 (원주 실측)
    자리 수를 세면 밀린다. 골든 표에 다가구 열은 없다."""
    ws = yb.sheet(r"주택", r"주택\s*현황")
    if ws is None:
        return None
    h = _header_row(ws, "연별") or 4
    hdr = list(range(h, h + 3))
    cols = {k: _find_col(ws, hdr, k) for k in
            ("일반가구", "합계", "단독주택", "아파트", "연립주택", "다세대", "비거주", "주택보급률")}
    yrs = _year_rows(ws)
    out = []
    for y in sorted(yrs)[-n:]:
        r = _row_values(ws, yrs[y])
        g = lambda k: _numc(r[cols[k] - 1]) if cols.get(k) else None
        out.append([y, g("일반가구"), g("합계"), g("단독주택"), g("아파트"),
                    g("연립주택"), g("다세대"), g("비거주"), g("주택보급률")])
    return out or None


def 사업체총괄(yb):
    """사업체 총괄 최신 연도 행 → {연도, 사업체: {계, 개인, 회사법인, 회사이외, 비법인},
    종사자: {…}}. 조직형태 블록마다 (사업체, 종사자) 열 쌍 — 사업체 열을 라벨로 찾고
    종사자는 그 오른쪽 칸이다.
    ⚠️ 최신 연도 라벨에 각주가 붙는다 (`20221)`) — `_year_rows_loose` 로 줍는다."""
    ws = yb.sheet(r"사업체", r"사업체\s*총괄")
    if ws is None:
        return None
    h = _header_row(ws, "합계") or 4
    hdr = list(range(h, h + 4))
    c계 = _find_col(ws, hdr, "사업체수")
    c종 = _find_col(ws, hdr, "종사자수")
    blocks = {k: _find_col(ws, hdr, k) for k in ("개인사업체", "회사법인", "회사이외", "비법인")}
    yrs = _year_rows_loose(ws)
    if not yrs or c계 is None:
        return None
    y = max(yrs)
    r = _row_values(ws, yrs[y])
    g = lambda c: _numc(r[c - 1]) if c else None
    return {"연도": y,
            "사업체": {"계": g(c계), "개인": g(blocks["개인사업체"]), "회사법인": g(blocks["회사법인"]),
                    "회사이외": g(blocks["회사이외"]), "비법인": g(blocks["비법인"])},
            "종사자": {"계": g(c종), "개인": g(blocks["개인사업체"] + 1 if blocks["개인사업체"] else None),
                    "회사법인": g(blocks["회사법인"] + 1 if blocks["회사법인"] else None),
                    "회사이외": g(blocks["회사이외"] + 1 if blocks["회사이외"] else None),
                    "비법인": g(blocks["비법인"] + 1 if blocks["비법인"] else None)}}


def extract_0500(path, 읍면=None):
    """0500 5.3 전부 → dict. 없는 것은 None (러프 원칙 — 지어내지 않는다)."""
    yb = YearBook(path)
    return {"인구추이": 인구_연별(yb), "읍면": (인구_읍면(yb, 읍면) if 읍면 else None),
            "인구동태": 인구동태(yb), "주택": 주택현황(yb), "사업체": 사업체총괄(yb)}


def self_test_0500(path=None):
    """원주 2024 기본통계(웹 재배포판) vs **원주 env-status 골든 5.3 값** 대조."""
    path = path or ROOT / "raw_data/web/stats_yearbook/원주/2024_엑셀"
    if not Path(path).exists():
        print(f"[skip] 원주 기본통계가 없습니다: {path}")
        return True
    r = extract_0500(path, 읍면="호저면")
    exp = {  # golden/small-env/원주_무장리/env-status.txt 5.3 (검증 단계에서만 연다)
        "인구추이[-1]": [2023, 171275, 366279, 181708, 184571, 2.1],
        "인구추이[0]": [2019, 154583, 352860, 175363, 177497, 2.3],
        "읍면": {"세대": 1849, "인구": 3527, "남": 1854, "여": 1673},
        "인구동태[-1]": [2023, 1934, 2644, 47132, 45778],
        "주택[-1]": [2023, 171190, 171463, 56445, 112208, 1690, 1120, None, 102.06],
        "사업체.계": (43665, 171895), "사업체.개인": (35145, 63220),
        "사업체.회사이외": (2017, 44931), "사업체.비법인": (978, 6744),
    }
    got = {
        "인구추이[-1]": r["인구추이"][-1], "인구추이[0]": r["인구추이"][0],
        "읍면": r["읍면"], "인구동태[-1]": r["인구동태"][-1], "주택[-1]": r["주택"][-1],
        "사업체.계": (r["사업체"]["사업체"]["계"], r["사업체"]["종사자"]["계"]),
        "사업체.개인": (r["사업체"]["사업체"]["개인"], r["사업체"]["종사자"]["개인"]),
        "사업체.회사이외": (r["사업체"]["사업체"]["회사이외"], r["사업체"]["종사자"]["회사이외"]),
        "사업체.비법인": (r["사업체"]["사업체"]["비법인"], r["사업체"]["종사자"]["비법인"]),
    }
    ok = True
    for k, e in exp.items():
        mark = "OK" if got[k] == e else "✗"
        if got[k] != e:
            ok = False
        print(f"  {mark} {k:<14} {got[k]}" + ("" if got[k] == e else f"  ← 기대 {e}"))
    print("0500 자체검증", "통과" if ok else "실패")
    return ok


def self_test(path=None):
    path = path or ROOT / "raw_data/nas/stats/천안_화덕리/천안_2023"
    if not Path(path).exists():
        print(f"[skip] 통계연보가 없습니다: {path}")
        print("       NAS 에서 내려받으세요 — catalog/review/stats_catalog.md 참조")
        return True
    yb = YearBook(path)
    print(f"편 {len(yb.volumes)}개 — {', '.join(v[:14] for v in yb.volumes[:4])} …\n")

    ok = True
    # 시군 전체 — 정답지는 2021년 행을 썼다 (최신 2022 가 아니다). 실증 §4-3
    got = land_use(yb, year=2021)
    for k, want in GOLDEN_CHEONAN["시군_2021"].items():
        hit = got.get(k)
        flag = "OK " if hit == want else "MISS"
        ok &= hit == want
        print(f"  [{flag}] 천안시 {k:<4} 기대 {want:>8,.2f} / 추출 {hit if hit is None else f'{hit:>8,.2f}'}")

    got = land_use(yb, region="동면")
    for k, want in GOLDEN_CHEONAN["동면"].items():
        hit = got.get(k)
        flag = "OK " if hit == want else "MISS"
        ok &= hit == want
        print(f"  [{flag}] 동면   {k:<4} 기대 {want:>8,.2f} / 추출 {hit if hit is None else f'{hit:>8,.2f}'}")

    got = vehicles(yb, year=2022)
    for k, want in GOLDEN_CHEONAN["자동차"].items():
        hit = got.get(k)
        flag = "OK " if hit == want else "MISS"
        ok &= hit == want
        print(f"  [{flag}] 자동차 {k:<5} 기대 {want:>8,} / 추출 {hit if hit is None else f'{hit:>8,}'}")

    got = zoning(yb, year=2021)
    for k, want in GOLDEN_CHEONAN["용도지역_2021"].items():
        hit = got.get(k)
        flag = "OK " if hit == want else "MISS"
        ok &= hit == want
        print(f"  [{flag}] 용도지역 {k:<6} 기대 {want:>8,.2f} / 추출 "
              f"{hit if hit is None else f'{hit:>8,.2f}'}")

    got = roads(yb, year=2022)
    for k, want in GOLDEN_CHEONAN["도로_2022"].items():
        hit = (got.get(k) or {}).get("개통연장")
        flag = "OK " if hit == want else "MISS"
        ok &= hit == want
        print(f"  [{flag}] 도로   {k:<6} 기대 {want:>9,} / 추출 "
              f"{hit if hit is None else f'{hit:>9,}'}")

    got = emitters(yb, year=2022)
    for k, want in GOLDEN_CHEONAN["배출_2022"].items():
        hit = got[k]["계"] if isinstance(got.get(k), dict) else got.get(k)
        flag = "OK " if hit == want else "MISS"
        ok &= hit == want
        print(f"  [{flag}] 배출   {k:<6} 기대 {want:>9,} / 추출 "
              f"{hit if hit is None else f'{hit:>9,}'}")

    got = heritage(yb, year=2022)
    for k, want in GOLDEN_CHEONAN["문화재_2022"].items():
        hit = got.get(k)
        flag = "OK " if hit == want else "MISS"
        ok &= hit == want
        print(f"  [{flag}] 문화재 {k:<6} 기대 {want:>9,} / 추출 "
              f"{hit if hit is None else f'{hit:>9,}'}")

    # 구성비 역산 — rule §3-1
    ratio = land_use_ratio(land_use(yb, year=2021))
    for k, want in (("임야", 48.58), ("전", 8.16), ("답", 14.60)):
        hit = ratio.get(k)
        flag = "OK " if hit == want else "MISS"
        ok &= hit == want
        print(f"  [{flag}] 구성비 {k:<4} 기대 {want:>8.2f}% / 계산 {hit:>8.2f}%")

    print("\n" + ("✅ 전 항목 일치" if ok else "❌ 불일치 있음"))
    return ok


def main():
    ap = argparse.ArgumentParser(description="통계연보 → 지역개황 표 값")
    ap.add_argument("path", nargs="?", help="통계연보 zip 또는 폴더")
    ap.add_argument("--region", help="읍면동 이름 (없으면 시군 전체)")
    ap.add_argument("--year", type=int, help="연도 행 (없으면 최신)")
    ap.add_argument("--what", default="land", choices=["land", "vehicle", "sheets"])
    ap.add_argument("--self-test", action="store_true")
    ap.add_argument("--self-test-0500", action="store_true",
                    help="0500 5.3 추출 vs 원주 env-status 골든")
    a = ap.parse_args()

    if a.self_test_0500:
        sys.exit(0 if self_test_0500(a.path) else 1)
    if a.self_test or not a.path:
        sys.exit(0 if self_test(a.path) else 1)

    yb = YearBook(a.path)
    if a.what == "sheets":
        for k, (vol, sh, mark) in SHEET_MAP.items():
            ws = yb.sheet(vol, sh)
            print(f"  {mark} {k:<26} → {ws.title if ws else '(못 찾음)'}")
    elif a.what == "land":
        area = land_use(yb, a.region, a.year)
        ratio = land_use_ratio(area)
        print(f"[{area.pop('_출처행')}]  단위 ㎢ / %")
        for k, v in area.items():
            print(f"  {k:<8} {v:>10,.2f}  {ratio.get(k, 0):>6.2f}%")
    else:
        for k, v in vehicles(yb, a.year).items():
            print(f"  {k:<8} {v}")


if __name__ == "__main__":
    main()
