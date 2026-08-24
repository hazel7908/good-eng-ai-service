#!/usr/bin/env python3
"""
엑셀 격자 읽기 — 정부 통계 엑셀의 머리글·병합·지자체 행을 다루는 공용 층.

전국 통계(상수도·하수도·폐기물…)는 기관마다 판마다 생김새가 다르다. 그런데
**깨지는 자리는 늘 같은 셋**이라 여기 모아 둔다.

    ① 머리글이 여러 줄이다        → header_band / header_names
    ② 병합 셀이 `None` 으로 온다   → 끌어 채우기 (단, 위층 경계에서 끊는다)
    ③ 열 번호가 판마다 밀린다      → find_col (이름으로 찾는다)

③ 은 실측이다 — 상수도통계 취수시설 시트에서 `일평균취수량` 이
**2021판 24열 · 2023판 26열**이었다. 이름은 세 판 모두 같았다.
→ `docs/20260824_지역개황_작업계획.md` §9-4

`engine/stats_national.py`(생성)와 `catalog/trace_stats.py`(역추적)가 같이 쓴다.
"""
import re

import openpyxl


def num(v):
    """숫자로 볼 수 있으면 float, 아니면 None."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.replace(",", "").replace(" ", "").strip()
        if re.fullmatch(r"-?\d+(\.\d+)?", s):
            return float(s)
    return None


def sheet_rows(path, sheet_pat):
    """시트를 이름 정규식으로 찾아 전 행을 리스트로 읽는다."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = [n for n in wb.sheetnames if re.search(sheet_pat, n)]
    if not names:
        raise KeyError(f"시트 없음: /{sheet_pat}/ — 있는 것: {wb.sheetnames[:8]}")
    return [list(r) for r in wb[names[0]].iter_rows(values_only=True)], names[0]


def header_band(rows, limit=14):
    """머리글 구간 — 숫자보다 글자가 많은 선두 행들."""
    band = []
    for r, row in enumerate(rows[:limit]):
        vals = [v for v in row if v is not None and str(v).strip()]
        if not vals:
            band.append(r)
            continue
        if sum(1 for v in vals if num(v) is None) / len(vals) < 0.5:
            break
        band.append(r)
    return band


def header_cols(rows, band):
    """열별 머리글을 **층 목록으로** 돌려준다 (`["6_공공하수처리시설", "시설용량(500㎥/일이상/미만)"]`).

    병합 셀은 `read_only` 에서 `None` 이라 왼쪽 값을 끌어 채워야 한다. 다만
    **무제한으로 끌면 아래층이 위층 경계를 넘어 옆 블록을 덮는다** — 상수도통계
    취수시설 시트에서 `해수`(수원형태별 세부)가 `일평균취수량` 자리까지 밀고 왔다.
    그래서 **위층이 바뀌는 지점에서 아래층 끌기를 끊는다.**

    ⚠️ 층을 문자열로 이어 붙이지 않는 이유: **머리글 자체에 `/` 가 들어간다**
    (`시설용량(500㎥/일이상/미만)`). 이어 붙이면 다시 못 가른다.
    """
    if not band:
        return []
    width = max((len(rows[r]) for r in band), default=0)
    cols = [[] for _ in range(width)]
    for r in band:
        row = list(rows[r]) + [None] * (width - len(rows[r]))
        last = last_parent = None
        vals = []
        for c in range(width):
            v = row[c]
            parent = "\x1f".join(cols[c])
            if v is not None and str(v).strip():
                last, last_parent = str(v).replace("\n", "").strip(), parent
            elif parent != last_parent:
                last = None          # 위층이 바뀌었다 — 여기서 끊는다
            vals.append(last)
        for c in range(width):
            if vals[c] and (not cols[c] or cols[c][-1] != vals[c]):
                cols[c].append(vals[c])
    return cols


def header_names(rows, band):
    """표시용 — 층을 `/` 로 이어 붙인다. **열 찾기에는 쓰지 말 것** (위 주의)."""
    return ["/".join(c[-3:]) for c in header_cols(rows, band)]


def find_col(cols, pat, required=True):
    """열 이름 정규식으로 열 번호를 찾는다 — 번호를 코드에 박지 않기 위해서.

    `cols` 는 `header_cols()` 의 **층 목록**이다. 정규식은 **각 층에 따로** 걸린다.
    `^취수장명$` 처럼 앵커를 쓰면 상위 층(`Ⅲ. 수도시설현황`)에 걸리지 않는다.
    아래 층이 더 구체적이므로 **뒤 층부터** 본다.
    """
    for i, parts in enumerate(cols):
        for seg in reversed(parts):
            if re.search(pat, seg or ""):
                return i
    if required:
        raise KeyError(f"열 없음: /{pat}/")
    return None


def region_rows(rows, start, region_col, region, skip_col=None,
                skip_values=("계", "소계", "합계")):
    """어느 지자체의 데이터 행만 고른다.

    ⚠️ 두 가지가 판마다 다르다.
      - 값이 `강원특별자치도 원주시` 처럼 **시도명이 앞에 붙는다** → 접미 일치로 본다
      - **2021판은 시도·시군 열이 병합**돼 아래 행이 `None` 이다 (2023판은 행마다 반복)
        → 마지막 값을 끌어 내린다
    """
    out, last = [], None
    for row in rows[start:]:
        if len(row) <= region_col:
            continue
        v = row[region_col]
        if v is not None and str(v).strip():
            last = str(v).strip()
        if not last:
            continue
        if not (last == region or last.endswith(" " + region)):
            continue
        if skip_col is not None and len(row) > skip_col:
            s = row[skip_col]
            if s is None or str(s).strip() in skip_values:
                continue
        out.append(row)
    return out
