#!/usr/bin/env python3
"""
**선언으로 표현되지 않는 모양**의 전국 통계.

`stats_national.py` 의 `SOURCES` 는 *시군 → 행 하나 = 시설 하나* 를 전제한다.
그 틀에 안 맞는 것들만 여기서 전용으로 읽는다 — 억지로 욱여넣지 않는다.

| 절 | 왜 다른가 | 형식 |
|---|---|---|
| 수변구역 | 시군 면적이 **한 셀 안 목록**에 있다 (`원주시(5.344)`) | hwpx |
| 생태경관보전지역 | 행=지역. 시군 열이 아예 없다 | hwpx |
| 하천일람 | **시군이 아니라 하천**으로 찾는다 | xlsx (시도별 시트) |
| 야생생물 보호구역 | **시군 열이 없다** — 보호지역명 문자열 안에 들어 있다 | xlsx |

    python engine/stats_irregular.py --self-test
    python engine/stats_irregular.py 수변구역 <파일.hwpx> --region 원주시
    python engine/stats_irregular.py 하천일람 <파일.xlsx> --region 섬강
"""
import argparse
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from hwpx_table import pick, tables                       # noqa: E402

CHECK = "[확인 필요]"
NAT = Path("raw_data/nas/stats/_national")


def 수변구역(path, region):
    """`4대강 수계 수변구역 지정현황` — 시군 면적이 **셀 안 목록**에 들어 있다.

    반환 {수계, 시도, 면적(㎢)} 또는 None(미지정).
    """
    t = pick(path, [r"수계구분", r"수변구역면적"])
    if not t:
        return None
    수계 = None
    for row in t:
        first = row[0] if row else ""
        if re.search(r"수계$", first):
            수계 = first
        for cell in row:
            # `원주시(5.344)` — 시군명 바로 뒤 괄호 안이 그 시군 면적이다
            m = re.search(rf"{re.escape(region)}\s*\(\s*([\d,.]+)\s*\)", cell)
            if m:
                시도 = next((c for c in row if re.search(r"(도|광역시|특별시)$", c)), CHECK)
                return {"수계": 수계 or CHECK, "시도": 시도,
                        "수변구역면적(㎢)": float(m.group(1).replace(",", ""))}
    return None


def 생태경관보전지역(path, region=None):
    """`생태·경관보전지역 지정현황` — 행 하나 = 지역 하나. 보통 표다.

    ⚠️ **면적이 판마다 다르다** — 동강유역 `'23.12월` 79.259 ↔ `'24.5월` 80.426
    (`24.3.8확대`). 폴리곤으로 계산하면 안 되는 이유다 (`ecgy.py` 주석 · F-10).
    """
    t = pick(path, [r"지역명|명\s*칭|구\s*분", r"면\s*적"]) or (tables(path) or [None])[-1]
    if not t:
        return []
    out = []
    for row in t:
        cells = [c for c in row if c]
        if len(cells) < 4:
            continue
        area = next((c for c in cells if re.fullmatch(r"[\d,]+\.\d+", c)), None)
        if not area:
            continue
        if region and not any(region.rstrip("시군구") in c for c in cells):
            continue
        out.append({"지역명": cells[0], "소재지": cells[1],
                    "면적(㎢)": float(area.replace(",", "")),
                    "특징": " / ".join(cells[cells.index(area) + 1:-1]) or CHECK,
                    "지정일자": cells[-1]})
    return out


def 수계_체인(path, 하천명, 시도=None):
    """사업지 인근 하천 → **최종 본류까지의 계통**. 수계흐름모식도·2.8.3 서술의 뼈대다.

    하천일람은 하천마다 `본류·제1지류·제2지류·제3지류` 를 갖고 있어 **한 줄이 곧 체인**이다.
    천안 정답(`용두천 → 병천천 → 미호천 → 금강`)의 병천천 행이 그대로 그 순서다.

    ⚠️ **동명이천이 있다** — `용두천` 은 세종에도 있고 그쪽 제1지류는 `대교천` 이다.
    이름만으로 찾으면 엉뚱한 수계가 나온다. **시도로 걸러야 한다.**

    ⚠️ **거리는 안 나온다.** 정답의 `2.97km · 9.19km` 는 지도에서 물길을 따라 잰 값이라
    자료에 없다 — 하천망이 면형이라 중심선을 못 뽑는다 (`hydro.py` 머리말).
    체인까지가 자료로 갈 수 있는 끝이다.
    """
    r = 하천일람(path, 하천명, 시도)
    if not r:
        return None
    # 제3 → 제2 → 제1 → 본류 순으로 위로 올라간다 (빈 칸은 건너뛴다)
    chain, seen = [], set()
    for k in ("제3지류", "제2지류", "제1지류", "본류"):
        nm = (r.get(k) or "").strip()
        if nm and nm not in seen:
            seen.add(nm)
            chain.append(nm)
    if 하천명 not in seen:
        chain.insert(0, 하천명)
    등급 = {}
    for nm in chain:
        g = 하천일람(path, nm, 시도) or 하천일람(path, nm)
        등급[nm] = (g or {}).get("하천등급", CHECK)
    return {"기준하천": 하천명, "체인": chain, "등급": 등급,
            "최종본류": chain[-1] if chain else CHECK,
            "거리": CHECK}          # 자료에 없다 — 지도 판독이 필요하다


def 하천일람(path, 하천명, 시도=None):
    """`한국하천일람` 시도별 — **시군이 아니라 하천으로 찾는다.**

    다른 절은 전부 `시군 → 행` 인데 이것만 다르다. 사업지가 어느 하천으로
    유하하는지는 **인풋(본문 서술)에서 오는 값**이라 여기서 정하지 않는다.

    반환 {하천명, 수계(본류·제1~3지류), 하천등급, 기점, 종점, 유로연장, 유역면적}.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = [시도] if 시도 else [n for n in wb.sheetnames if len(n) <= 3]
    for name in sheets:
        if name not in wb.sheetnames:
            continue
        for row in wb[name].iter_rows(values_only=True):
            if not row or not row[0] or str(row[0]).strip() != 하천명:
                continue
            g = lambda i: (str(row[i]).replace("\n", " ").strip()
                           if i < len(row) and row[i] is not None else "")
            return {
                "하천명": 하천명, "시도": name,
                "본류": g(1), "제1지류": g(2), "제2지류": g(3), "제3지류": g(4),
                "하천등급": g(9),
                "기점": g(12), "종점": g(13),
                "유로연장(㎞)": g(14), "유역면적(㎢)": g(15),
            }
    return None


def 야생생물보호구역(path, region):
    """`야생생물 보호구역 현황` — **시군 열이 없다.**

    보호지역명이 `강원 원주 소초면(원주시 소초면 학곡리 산 56번지외 1필지)` 꼴이라
    **이름 안에서 시군을 찾는다.** 골든셋 표는 `연번 · 소재지 · 면적(㎢) · 비고` 다.

    ⚠️ 창고 최신이 **2017년 12월 기준**이다 (골든셋도 그 판을 인용한다).
    더 새 판은 발행처를 뚫어야 한다 — 아직 경로 미확인.
    """
    import openpyxl
    key = region.rstrip("시군구")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        name = row[2] if len(row) > 2 else None
        if not name or key not in str(name):
            continue
        out.append({"연번": row[0], "소재지": str(name).strip(),
                    "면적(㎢)": row[6] if len(row) > 6 else CHECK, "비고": "-"})
    return out


GOLDEN = [
    ("원주 수변구역 2024", "수변구역", NAT / "4대강 수계 수변구역 지정현황(''24년 기준).hwpx",
     "원주시", {"수변구역면적(㎢)": 5.344, "수계": "한강수계"}),
    ("원주 하천일람 2022 (섬강)", "하천일람",
     NAT / "2022년 한국하천일람(시도별하천일람_2022.12.31 기준).xlsx",
     "섬강", {"본류": "한강", "제1지류": "섬강", "하천등급": "국가"}),
    ("원주 야생생물 2017", "야생생물보호구역",
     NAT / "야생생물보호구역 현황(171231).xlsx", "원주시",
     {"연번": 11, "면적(㎢)": 0.059}),
    ("평창 생태경관 '23.12월", "생태경관보전지역",
     NAT / "생태경관보전지역 지정현황('23.12월).hwpx", "평창", {"면적(㎢)": 79.259}),
]


def self_test():
    ok = bad = 0
    for label, fn, path, region, expect in GOLDEN:
        if not path.exists():
            print(f"⏭  {label}: 원자료 없음"); continue
        got = globals()[fn](path, region)
        if isinstance(got, list):
            got = (next((g for g in got if "동강" in g.get("지역명", "")), None)
                   or (got[0] if got else None))
        print(f"\n== {label}")
        if not got:
            print("   ❌ 값 없음"); bad += 1; continue
        for k, e in expect.items():
            g = got.get(k)
            hit = abs(g - e) < 0.001 if isinstance(e, float) else str(g) == str(e)
            print(f"   {'✅' if hit else '❌'} {k}: {g!r}" + ("" if hit else f" ≠ {e!r}"))
            ok, bad = (ok + 1, bad) if hit else (ok, bad + 1)
    print(f"\n=== 자체검증: {ok} OK · {bad} 불일치 ===")
    return 1 if bad else 0


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("절", nargs="?",
                    choices=["수변구역", "생태경관보전지역", "하천일람",
                             "야생생물보호구역", "수계_체인"])
    ap.add_argument("hwpx", nargs="?")
    ap.add_argument("--region")
    ap.add_argument("--self-test", action="store_true")
    a = ap.parse_args()
    if a.self_test:
        return self_test()
    print(globals()[a.절](a.hwpx, a.region))
    return 0


if __name__ == "__main__":
    sys.exit(main())
