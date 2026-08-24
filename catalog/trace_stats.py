#!/usr/bin/env python3
"""
통계 원자료 역추적기 — 보고서의 값이 원자료 **어디서** 왔는지 찾는다.

지역개황의 통계 표는 사람이 엑셀에서 눈으로 찾아 옮긴 것이다.
그 "어느 시트 어느 열" 을 우리가 다시 손으로 찾으면 같은 실수를 반복한다.
**값을 넣으면 자료가 자리를 알려준다** — 좌표를 눈대중이 아니라 증거로 만든다.

    python catalog/trace_stats.py <엑셀> --values 15600 12460.45 1300
    python catalog/trace_stats.py <엑셀> --values 원주 흥업 --near 원주시
    python catalog/trace_stats.py <엑셀> --golden golden/…/regional-overview.txt --sec 2.7.1

출력은 값마다 (시트 · 행 · 열 · 추정 열이름). 여러 시트에서 걸리면 전부 보여준다 —
**어느 것이 맞는지는 값 여러 개가 같은 시트에 몰리는가로 판단한다.**

⚠️ 이 도구는 **검증·분석 전용**이다. 생성 파이프라인에 들어가지 않는다.
"""
import argparse
import re
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "engine"))
from xlsx_grid import header_band, header_cols, header_names, num as _num


def _norm(s):
    """비교용 문자열 정규화 — 공백·괄호·중점 제거."""
    if s is None:
        return ""
    return re.sub(r"[\s()（）·ㆍ,]", "", str(s))


def scan(path, targets, max_hits=20):
    """targets(값 목록)를 전 시트에서 찾는다. 반환 {값: [(시트,행,열,헤더), …]}."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    nums = {t: _num(t) for t in targets}
    txts = {t: _norm(t) for t in targets}
    found = {t: [] for t in targets}

    for name in wb.sheetnames:
        ws = wb[name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        hdr = header_names(rows, header_band(rows))
        for r, row in enumerate(rows):
            for c, cell in enumerate(row):
                if cell is None:
                    continue
                cn, ct = _num(cell), _norm(cell)
                for t in targets:
                    if len(found[t]) >= max_hits:
                        continue
                    tn = nums[t]
                    hit = False
                    if tn is not None and cn is not None:
                        hit = abs(cn - tn) < 1e-6 or (
                            abs(tn) > 1 and abs(cn - tn) / abs(tn) < 1e-9
                        )
                    elif tn is None and ct and txts[t]:
                        hit = txts[t] == ct or (len(txts[t]) >= 2 and txts[t] in ct)
                    if hit:
                        found[t].append((name, r + 1, c, hdr[c] if c < len(hdr) else ""))
    return found


def report(found, path):
    print(f"# 역추적: {Path(path).name}\n")
    sheet_score = {}
    for t, hits in found.items():
        mark = "✅" if hits else "❌"
        print(f"{mark} {t}")
        for s, r, c, h in hits:
            print(f"      {s}  행{r} 열{c}  ← {h}")
            sheet_score[s] = sheet_score.get(s, 0) + 1
        if not hits:
            print("      (원자료 어디에도 없다)")
    if sheet_score:
        print("\n## 시트별 적중 수 — 많이 몰린 곳이 그 표의 출처다")
        for s, n in sorted(sheet_score.items(), key=lambda x: -x[1]):
            print(f"  {n:3d}  {s}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("excel")
    ap.add_argument("--values", nargs="+", required=True, help="찾을 값(숫자/문자열)")
    ap.add_argument("--max-hits", type=int, default=20)
    a = ap.parse_args()
    report(scan(a.excel, a.values, a.max_hits), a.excel)


if __name__ == "__main__":
    sys.exit(main())
