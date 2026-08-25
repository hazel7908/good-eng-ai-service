#!/usr/bin/env python3
"""매립 원자료 진단 — 천안 2행 문제의 뿌리를 한 번에 찾는다 (2026-08-25).

천안 지역개황에서 매립처리시설이 **2행**으로 나왔는데 정답은 **1행**(공공)이다.
그리고 정답의 기매립량 `66,026.2` 는 우리 값 `1,372,592.5` 의 1/20 이다.
둘 중 무엇이 원인인지 원자료를 봐야 갈린다.

    python catalog/probe_landfill.py            # 자료 자동 탐색
    python catalog/probe_landfill.py <엑셀경로>

⚠️ `raw_data/` 는 git 제외라 **맥에서만 돌아간다.**
⚠️ 검증·분석 전용이다. 생성 파이프라인에 들어가지 않는다.

읽는 법
  ① `천안제3산단` 이 **공공매립 시트에 있으면** → 이름이 아니라 열로 걸러야 한다.
     시트에 없으면 → 우리 시트 정규식이 다른 시트까지 먹은 것이다 (코드 버그).
  ② `66026.2` 가 **어디에도 없으면** → 정답지 오류가 확정된다.
     어느 열엔가 있으면 → 우리 열 매핑이 틀린 것이다.
"""
import re
import sys
from pathlib import Path

import openpyxl

NAT = Path("raw_data/nas/stats/_national")
TARGET = ["천안제3산단", "목천위생", "66026.2", "1372592.5"]


def find_book(argv):
    if len(argv) > 1:
        return Path(argv[1])
    hits = sorted(NAT.rglob("*처리업체현황*.xlsx"))
    if not hits:
        sys.exit(f"자료를 못 찾았다 — {NAT}/**/*처리업체현황*.xlsx\n"
                 f"경로를 인자로 넘겨라: python {argv[0]} <엑셀>")
    if len(hits) > 1:
        print(f"⚠️ 후보 {len(hits)}개 — 첫 번째를 쓴다")
        for h in hits:
            print(f"   {h}")
    return hits[0]


def cell_text(c):
    return "" if c is None else str(c).strip()


def main():
    book = find_book(sys.argv)
    print(f"자료: {book}\n")
    wb = openpyxl.load_workbook(book, read_only=True, data_only=True)
    print(f"시트 {len(wb.sheetnames)}개")
    매립시트 = [s for s in wb.sheetnames if "매립" in s]
    print(f"  '매립' 들어간 시트: {매립시트}")
    print(f"  우리 정규식 r'공공매립' 이 먹는 시트: "
          f"{[s for s in wb.sheetnames if re.search('공공매립', s)]}\n")

    for sn in 매립시트:
        ws = wb[sn]
        rows = [[cell_text(c) for c in r] for r in ws.iter_rows(values_only=True)]
        print(f"══ [{sn}] {len(rows)}행")
        # 머리행 — '시설명' 이 든 첫 행부터 두 줄
        hi = next((i for i, r in enumerate(rows[:20]) if any("시설명" in c for c in r)), 0)
        for i in range(hi, min(hi + 3, len(rows))):
            print(f"   머리{i}: {[c for c in rows[i] if c][:14]}")
        # 공공/민간 구분 열이 있는가
        band = " ".join(c for i in range(hi, min(hi + 3, len(rows))) for c in rows[i])
        print(f"   ▸ 공공/민간 구분 열: {'있다 ✅' if re.search('공공|민간|구 ?분', band) else '없다'}")
        # 천안 행
        천안 = [r for r in rows if any("천안" in c for c in r)]
        print(f"   ▸ 천안 행 {len(천안)}개")
        for r in 천안:
            print(f"      {[c for c in r if c][:10]}")
        print()

    print("══ 값 역추적 (전 시트)")
    for t in TARGET:
        hits = []
        for sn in wb.sheetnames:
            for i, r in enumerate(wb[sn].iter_rows(values_only=True)):
                for j, c in enumerate(r):
                    s = cell_text(c).replace(",", "")
                    if s == t or (t in s and not t.replace(".", "").isdigit()):
                        hits.append(f"{sn}!R{i+1}C{j+1}")
                        break
                if len(hits) > 6:
                    break
        print(f"   {t:14} {'· '.join(hits) if hits else '❌ 어디에도 없다'}")


if __name__ == "__main__":
    main()
